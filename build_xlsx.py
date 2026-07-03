from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

ORANGE="F84001"; GREEN="024632"; GMID="13BC65"; AMBER="FE930D"; RED="E52F2D"
BEIGE="F8E0C5"; INK="322F2E"; BROWN="6F5F4F"; WHITE="FFFFFF"
F="Calibri"
def font(b=False,c=INK,sz=11,it=False): return Font(name=F,bold=b,color=c,size=sz,italic=it)
def fill(c): return PatternFill("solid",start_color=c)
thin=Side(style="thin",color="D9D2C8")
border=Border(left=thin,right=thin,top=thin,bottom=thin)
center=Alignment(horizontal="center",vertical="center",wrap_text=True)
left=Alignment(horizontal="left",vertical="center",wrap_text=True)

wb=Workbook()

# ---------------- LEIA-ME ----------------
ws=wb.active; ws.title="Leia-me"; ws.sheet_view.showGridLines=False
ws.column_dimensions['A'].width=2; ws.column_dimensions['B'].width=110
ws['B2']="refuturiza 360"; ws['B2'].font=Font(name=F,bold=True,color=ORANGE,size=22)
ws['B3']="Sistema de Acompanhamento de Customer Success"; ws['B3'].font=font(True,GREEN,15)
ws['B4']="Emprego e Educação para Todos"; ws['B4'].font=font(False,BROWN,9)
rows=[
 ("",""),
 ("COMO USAR","h"),
 ("1. Vá na aba CONTAS e preencha uma linha por cliente. Só edite as colunas AZUIS (entrada).","p"),
 ("2. As colunas LARANJAS são calculadas automaticamente: uso de licenças, sub-scores, Health Score, Risco de Churn, Farol, Upgrade.","p"),
 ("3. A aba PAINEL mostra os indicadores agregados (NRR, GRR, retenção de logos, ARR em risco, expansão potencial).","p"),
 ("4. A aba METODOLOGIA explica cada cálculo, as premissas e os pontos cegos de CS.","p"),
 ("",""),
 ("PARA ALIMENTAR O DASHBOARD","h"),
 ("Salve a aba CONTAS como CSV (Arquivo > Salvar como > CSV) e importe no dashboard HTML (botão \"Importar CSV\").","p"),
 ("",""),
 ("VERDADE IMPORTANTE","h"),
 ("O Risco de Churn aqui é um PROXY baseado em sinais de CRM/relacionamento/suporte — não usa dados de uso do produto (logins, features).","p"),
 ("É ótimo para priorizar atenção, mas não é um modelo preditivo validado. Para isso, conecte uma fonte de uso (ex: Amplitude) e valide o score contra churn real.","p"),
 ("",""),
 ("As 12 linhas já preenchidas são EXEMPLOS fictícios. Substitua pelos seus clientes.","i"),
]
r=5
for txt,sty in rows:
    c=ws.cell(row=r,column=2,value=txt)
    if sty=="h": c.font=font(True,ORANGE,13)
    elif sty=="i": c.font=font(False,BROWN,10,it=True)
    else: c.font=font(False,INK,11)
    c.alignment=left
    r+=1

# ---------------- CONTAS ----------------
cs=wb.create_sheet("Contas"); cs.sheet_view.showGridLines=False
inputs=[("Conta",22),("CSM",12),("Plano",13),("MRR inicial\n(12m atrás)",13),("MRR atual",12),
        ("Licenças\ncontratadas",12),("Licenças\nativas",11),("Freq. login\n(0-100)",11),
        ("Dias s/\ncontato",10),("Champion?",11),("Nº contatos\nengajados",11),
        ("Tickets\nabertos",10),("Tickets\ncríticos",10),("CSAT\n(0-100)",9),("NPS\n(-100..100)",11),
        ("Pagamento",12),("Estágio jornada",15),("Dias p/\nrenovação",11)]
calcs=[("Uso\nlicenças %",11),("Sub\nAdoção",9),("Sub\nRelac.",9),("Sub\nSuporte",9),
       ("Sub\nComercial",10),("Sub\nSentim.",9),("HEALTH\nSCORE",11),("RISCO\nCHURN",10),
       ("FAROL",12),("Upgrade?",10),("Expansão\npot. (R$/mês)",13)]
inputs=[(h.replace(chr(10),' '),w) for h,w in inputs]
calcs=[(h.replace(chr(10),' '),w) for h,w in calcs]
headers=inputs+calcs
ncols=len(headers)
# title
cs.cell(row=1,column=1,value="CARTEIRA DE CLIENTES — preencha as colunas AZUIS").font=font(True,WHITE,12)
cs.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncols)
cs.cell(row=1,column=1).fill=fill(GREEN); cs.cell(row=1,column=1).alignment=left
# headers row 2
for j,(h,w) in enumerate(headers,1):
    c=cs.cell(row=2,column=j,value=h)
    is_input=j<=len(inputs)
    c.fill=fill(GREEN if is_input else ORANGE)
    c.font=font(True,WHITE,9); c.alignment=center; c.border=border
    cs.column_dimensions[get_column_letter(j)].width=w
cs.row_dimensions[2].height=30
cs.freeze_panes="A3"

sample=[
 ["Construtora Horizonte","Marina","Enterprise",32000,38000,500,430,82,9,"Sim",6,2,0,88,9,"Em dia","Expansão",120],
 ["Rede Bem Estar Saúde","Rafael","Growth",18000,18000,300,96,31,54,"Não",1,5,1,61,2,"Em dia","Engajamento",160],
 ["Grupo Varejo Sul","Marina","Enterprise",41000,41000,800,740,77,14,"Sim",8,3,0,84,8,"Em dia","Renovação",35],
 ["TechNova Sistemas","Bruno","Growth",12000,9000,200,70,24,71,"Não",1,7,2,48,-15,"Atrasado","Renovação",35],
 ["Indústria Ferro Forte","Rafael","Starter",5000,6500,80,74,69,18,"Sim",4,1,0,80,7,"Em dia","Adoção",240],
 ["Logística Andorinha","Bruno","Growth",16000,16000,250,235,74,22,"Sim",5,2,0,79,6,"Em dia","Expansão",120],
 ["Banco Coop Aliança","Marina","Enterprise",55000,48000,1000,520,41,38,"Não",2,6,1,58,-5,"Em dia","Renovação",35],
 ["Rede Educa Mais","Rafael","Growth",14000,15500,220,198,73,11,"Sim",5,1,0,86,9,"Em dia","Advocacia",200],
 ["Atacado Real Distrib.","Bruno","Starter",4000,4000,60,22,28,63,"Não",1,3,0,55,0,"Atrasado","Onboarding",300],
 ["Hospital Vida Plena","Marina","Enterprise",38000,44000,700,610,80,7,"Sim",7,2,0,90,10,"Em dia","Expansão",120],
 ["Transportes Litoral","Rafael","Growth",13000,13000,180,120,52,29,"Sim",3,4,0,70,4,"Em dia","Engajamento",160],
 ["AgroVerde Cooperativa","Bruno","Starter",6000,7000,90,81,66,16,"Sim",4,0,0,82,8,"Em dia","Adoção",240],
]
NROWS=120  # capacity
blue=font(False,"0000FF",10)
black=font(False,INK,10)
for i in range(NROWS):
    r=3+i
    rowdata=sample[i] if i<len(sample) else [None]*len(inputs)
    for j in range(len(inputs)):
        c=cs.cell(row=r,column=j+1,value=rowdata[j])
        c.font=blue; c.border=border
        c.alignment=left if j in (0,1,2,9,15,16) else Alignment(horizontal="center",vertical="center")
    # calc columns formulas
    S=f"R{r}"  # placeholder not used
    # column letters
    A=get_column_letter; 
    # input col letters: D MRRini E MRRnow F entit G act H login I days J champ K contacts L tick M crit N csat O nps P pay Q stage R renew
    f_use=f'=IF(F{r}=0,0,G{r}/F{r})'
    f_ado=f'=MEDIAN(0,0.6*(S{r}*100)+0.4*H{r},100)'
    f_rel=f'=MEDIAN(0,IF(J{r}="Sim",45,0)+MEDIAN(0,K{r}*10,30)+MEDIAN(0,25-I{r}/3,25),100)'
    f_sup=f'=MEDIAN(0,0.5*N{r}+0.3*MEDIAN(0,100-L{r}*12,100)+0.2*MEDIAN(0,100-M{r}*40,100),100)'
    f_com=f'=MEDIAN(0,IF(P{r}="Em dia",60,20)+IF(E{r}>=D{r},40,IF(E{r}>=D{r}*0.85,20,0)),100)'
    f_sen=f'=MEDIAN(0,(O{r}+100)/2,100)'
    f_health=f'=ROUND(0.35*T{r}+0.20*U{r}+0.20*V{r}+0.15*W{r}+0.10*X{r},0)'
    f_churn=(f'=ROUND(MEDIAN(0,(1-S{r})*25+IF(H{r}<40,15,IF(H{r}<60,7,0))+IF(J{r}="Sim",0,18)'
             f'+IF(K{r}<=1,10,0)+MEDIAN(0,I{r}/4,18)+M{r}*8+IF(E{r}<D{r},12,0)'
             f'+IF(P{r}="Atrasado",10,0)+IF(O{r}<0,8,0),100),0)')
    f_farol=(f'=IF(N{r}="","",IF(OR(E{r}=0,Y{r}<45,Z{r}>=65,AND(R{r}<=60,Y{r}<60,E{r}>0)),"Vermelho",'
             f'IF(OR(AND(Y{r}>=45,Y{r}<70),AND(Z{r}>=45,Z{r}<65)),"Amarelo","Verde")))')
    f_upg=f'=IF(AND(E{r}>0,OR(S{r}>=0.85,AND(Y{r}>=75,O{r}>=8,E{r}>=D{r},S{r}>=0.7))),"Sim","Não")'
    f_exp=f'=IF(AB{r}="Sim",ROUND(E{r}*0.25,0),0)'
    formulas=[f_use,f_ado,f_rel,f_sup,f_com,f_sen,f_health,f_churn,f_farol,f_upg,f_exp]
    # guard: only compute when Conta present -> wrap each to IF(A blank,"")
    for k,formula in enumerate(formulas):
        col=len(inputs)+1+k
        # wrap with blank guard except farol already guarded
        if k!=8:
            formula=f'=IF($A{r}="","",{formula[1:]})'
        c=cs.cell(row=r,column=col,value=formula)
        c.font=black; c.border=border
        c.alignment=Alignment(horizontal="center",vertical="center")
        if k==6: c.font=font(True,INK,10)
        if k==7: c.font=font(True,INK,10)
        if k in (0,): c.number_format='0%'
        if k==10: c.number_format='#,##0'
    # zebra
    if i%2==1:
        for j in range(1,ncols+1):
            cs.cell(row=r,column=j).fill=fill("FBF6EF")

# data validations
dv_ch=DataValidation(type="list",formula1='"Sim,Não"',allow_blank=True)
dv_pay=DataValidation(type="list",formula1='"Em dia,Atrasado"',allow_blank=True)
dv_st=DataValidation(type="list",formula1='"Onboarding,Adoção,Engajamento,Expansão,Renovação,Advocacia,Churn"',allow_blank=True)
cs.add_data_validation(dv_ch); cs.add_data_validation(dv_pay); cs.add_data_validation(dv_st)
last=2+NROWS
dv_ch.add(f"J3:J{last}"); dv_pay.add(f"P3:P{last}"); dv_st.add(f"Q3:Q{last}")

# conditional formatting on Farol col (AA = col 27)
far=f"AA3:AA{last}"
cs.conditional_formatting.add(far,CellIsRule(operator="equal",formula=['"Vermelho"'],fill=fill(RED),font=font(True,WHITE,10)))
cs.conditional_formatting.add(far,CellIsRule(operator="equal",formula=['"Amarelo"'],fill=fill(AMBER),font=font(True,WHITE,10)))
cs.conditional_formatting.add(far,CellIsRule(operator="equal",formula=['"Verde"'],fill=fill(GMID),font=font(True,WHITE,10)))

# ---------------- PAINEL ----------------
ps=wb.create_sheet("Painel"); ps.sheet_view.showGridLines=False
ps.column_dimensions['A'].width=2
for col,w in zip("BCDE",[34,18,4,46]): ps.column_dimensions[col].width=w
ps['B2']="PAINEL DE CUSTOMER SUCCESS"; ps['B2'].font=Font(name=F,bold=True,color=ORANGE,size=18)
ps['B3']="Indicadores agregados — recalculados a partir da aba Contas"; ps['B3'].font=font(False,BROWN,10)
D=f"D3:D{last}"; E=f"E3:E{last}"; Y=f"Y3:Y{last}"; AA=f"AA3:AA{last}"; AB=f"AB3:AB{last}"; AC=f"AC3:AC{last}"
metrics=[
 ("NRR — Net Revenue Retention",f'=IF(SUM(Contas!{D})=0,0,SUM(Contas!{E})/SUM(Contas!{D}))','0.0%',"Métrica nº 1 do B2B SaaS. >100% = base cresce mesmo perdendo clientes."),
 ("GRR — Gross Revenue Retention",f'=IF(SUM(Contas!{D})=0,0,SUMPRODUCT((Contas!{E}<Contas!{D})*Contas!{E}+(Contas!{E}>=Contas!{D})*Contas!{D})/SUM(Contas!{D}))','0.0%',"Retenção sem crédito de expansão. Mede vazamento puro."),
 ("Retenção de logos",f'=IF(COUNTIF(Contas!{D},">0")=0,0,COUNTIFS(Contas!{E},">0",Contas!{D},">0")/COUNTIF(Contas!{D},">0"))','0.0%',"% de contas que continuaram (não churnaram)."),
 ("ARR ativo total",f'=SUMIF(Contas!{E},">0")*12','"R$ "#,##0',"Receita recorrente anual das contas ativas."),
 ("Expansion MRR",f'=SUMPRODUCT((Contas!{E}>0)*((Contas!{E}-Contas!{D})>0)*(Contas!{E}-Contas!{D}))','"R$ "#,##0',"Soma dos aumentos de MRR nas contas retidas."),
 ("ARR em risco (farol vermelho)",f'=SUMIFS(Contas!{E},Contas!{AA},"Vermelho",Contas!{E},">0")*12','"R$ "#,##0',"Receita anual exposta em contas de risco."),
 ("Health Score médio (ativos)",f'=IFERROR(AVERAGEIFS(Contas!{Y},Contas!{E},">0"),0)','0',"Saúde média da carteira ativa (0-100)."),
 ("Contas saudáveis (verde)",f'=COUNTIFS(Contas!{AA},"Verde",Contas!{E},">0")','0',""),
 ("Contas em atenção (amarelo)",f'=COUNTIFS(Contas!{AA},"Amarelo",Contas!{E},">0")','0',""),
 ("Contas em risco (vermelho)",f'=COUNTIFS(Contas!{AA},"Vermelho",Contas!{E},">0")','0',""),
 ("Oportunidades de upgrade",f'=COUNTIF(Contas!{AB},"Sim")','0',"Contas prontas para expansão/upsell."),
 ("Expansão potencial (anual)",f'=SUM(Contas!{AC})*12','"R$ "#,##0',"Receita anual adicional estimada se concretizar upgrades."),
]
r=5
for name,formula,fmt,note in metrics:
    ps.cell(row=r,column=2,value=name).font=font(True,INK,11); ps.cell(row=r,column=2).alignment=left
    c=ps.cell(row=r,column=3,value=formula); c.number_format=fmt; c.font=font(True,GREEN,12); c.alignment=center
    c.fill=fill(BEIGE); c.border=border
    ps.cell(row=r,column=5,value=note).font=font(False,BROWN,9); ps.cell(row=r,column=5).alignment=left
    r+=1

# ---------------- METODOLOGIA ----------------
ms=wb.create_sheet("Metodologia"); ms.sheet_view.showGridLines=False
ms.column_dimensions['A'].width=2; ms.column_dimensions['B'].width=115
meth=[
 ("Metodologia, premissas e pontos cegos","t"),
 ("",""),
 ("HEALTH SCORE (0-100) — média ponderada de 5 dimensões:","h"),
 ("• Adoção/uso (35%): uso de licenças (ativas/contratadas) + frequência de login","p"),
 ("• Relacionamento (20%): champion, nº de contatos engajados (multithreading), dias sem contato","p"),
 ("• Suporte (20%): tickets abertos, tickets críticos, CSAT","p"),
 ("• Comercial (15%): status de pagamento, tendência de MRR","p"),
 ("• Sentimento (10%): NPS","p"),
 ("Os pesos são ponto de partida de mercado — recalibre contra seu churn real.","i"),
 ("",""),
 ("FAROL DE RISCO (a regra mais severa vence):","h"),
 ("• Vermelho: health < 45  OU  churn >= 65  OU  (renovação <= 60 dias E health < 60)","p"),
 ("• Amarelo: health 45-69  OU  churn 45-64","p"),
 ("• Verde: health >= 70  E  churn < 45","p"),
 ("",""),
 ("RISCO DE CHURN (proxy 0-100, maior = pior):","h"),
 ("Soma de sinais negativos: baixa adoção, uso baixo, sem champion, single-thread, dias sem contato,","p"),
 ("tickets críticos, MRR em queda, pagamento atrasado e NPS detrator.","p"),
 ("",""),
 ("UPGRADE/EXPANSÃO:","h"),
 ("Uso de licenças >= 85%  OU  (health >= 75 + NPS >= 8 + MRR estável/crescente + uso >= 70%).","p"),
 ("",""),
 ("PONTOS CEGOS QUE VOCÊ PEDIU PARA EU APONTAR:","h"),
 ("1. NRR > churn. A métrica que define um B2B SaaS é a Net Revenue Retention, não evitar perdas.","p"),
 ("2. Indicador atrasado vs. antecedente. Churn/renovação são lagging; o valor está nos sinais antecedentes (queda de uso, champion que saiu).","p"),
 ("3. Time-to-value/onboarding é o preditor mais forte de retenção — meça tempo até o 1º valor e % de onboarding concluído.","p"),
 ("4. Single-threading: conta dependente de 1 champion é frágil. Acompanhe nº de contatos engajados.","p"),
 ("5. Churn silencioso: contas que não renovam sem reclamar — só pegas por queda de uso (exige dado de produto).","p"),
 ("6. Segmentação/cost-to-serve: separe high-touch de tech-touch; avalie ARR por CSM.","p"),
 ("7. Validação do score: compare contas que churnaram — estavam vermelhas antes? Se não, recalibre.","p"),
 ("",""),
 ("LIMITAÇÃO PRINCIPAL: sem dados de uso do produto, o churn é um PROXY de CRM — útil, mas não preditivo validado.","i"),
]
r=2
for txt,sty in meth:
    c=ms.cell(row=r,column=2,value=txt)
    if sty=="t": c.font=Font(name=F,bold=True,color=ORANGE,size=16)
    elif sty=="h": c.font=font(True,GREEN,12)
    elif sty=="i": c.font=font(False,BROWN,10,it=True)
    else: c.font=font(False,INK,11)
    c.alignment=left
    r+=1

wb.save("CS-Refuturiza-360.xlsx")
print("saved to CS-Refuturiza-360.xlsx")
